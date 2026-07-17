"""Core Specifics: 30_Konfirmasi, 31_GC, 32_SaldoAwal, 33_PeristiwaSetelah, 34_PihakBerelasi,
35_GrupAudit, 36_Reliance, 37_SAD, 38_EvaluasiBukti."""
from xhelp import *
from openpyxl.worksheet.datavalidation import DataValidation
from gen_ref_setup import sum_cats, MAP_AL, MAP_LJPDK, MAP_LR_PRATAX_L, MAP_LIAB_L, MAP_EK_LR_L

WTB_J = "'20_WTB'!$J$6:$J$155"
WTB_C = "'20_WTB'!$C$6:$C$155"


def dv_inline(ws, items, rng):
    dv = DataValidation(type='list', formula1=f'"{",".join(items)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def build_konfirmasi(wb):
    ws = wb.create_sheet('30_Konfirmasi')
    title(ws, '30 · Confirmation Hub (SA 505)', 'Log pengendalian konfirmasi eksternal. Selisih otomatis; alternatif wajib bila tidak kembali.')
    widths(ws, {'A': 6, 'B': 26, 'C': 14, 'D': 10, 'E': 16, 'F': 12, 'G': 12, 'H': 12, 'I': 16, 'J': 16, 'K': 16, 'L': 34})
    cell(ws, 'A3', 'Terkirim:', font=F_BOLD, border=False)
    frm(ws, 'B3', '=COUNTA(F6:F45)', fmt=NUM, bold=True)
    cell(ws, 'C3', 'Kembali:', font=F_BOLD, border=False)
    frm(ws, 'D3', '=COUNTIF(I6:I45,"Cocok")+COUNTIF(I6:I45,"Selisih")', fmt=NUM, bold=True)
    cell(ws, 'E3', 'Coverage saldo terkonfirmasi:', font=F_BOLD, border=False)
    frm(ws, 'G3', '=IF(SUM(E6:E45)=0,"",SUMIFS(E6:E45,I6:I45,"Cocok")/SUM(E6:E45))', fmt=PCT, bold=True)
    headrow(ws, 5, ['No', 'Pihak Dikonfirmasi', 'Jenis', 'Ref WP', 'Saldo per Buku', 'Kirim ke-1', 'Kirim ke-2',
                    'Diterima', 'Hasil', 'Saldo per Konfirmasi', 'Selisih', 'Prosedur Alternatif / Catatan'])
    for r in range(6, 46):
        for colref in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L']:
            inp(ws, f'{colref}{r}', fmt=RP if colref in 'EJ' else (DATE if colref in 'FGH' else None))
        frm(ws, f'K{r}', f'=IF(OR($E{r}="",$J{r}=""),"",$E{r}-$J{r})', fmt=RP)
    dv_name(ws, 'LIST_JENIS_KONF', 'C6:C45')
    dv_name(ws, 'LIST_HASIL_KONF', 'I6:I45')
    ws['A6'] = 1; ws['B6'] = 'Bank ABC (contoh)'; ws['C6'] = 'Bank'; ws['E6'] = 500_000_000
    ws.freeze_panes = 'A6'
    protect(ws)
    return ws


def build_gc(wb):
    ws = wb.create_sheet('31_GC')
    title(ws, '31 · Going Concern (SA 570)', 'Rasio ditarik live dari WTB. Indikator dievaluasi; kesimpulan menentukan dampak ke opini.')
    widths(ws, {'A': 46, 'B': 16, 'C': 14, 'D': 40})
    cell(ws, 'A4', 'A. RASIO KEUANGAN (otomatis)', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 5, ['Rasio', 'Nilai', 'Benchmark', 'Indikasi'])
    rasios = [
        ('Rasio lancar (aset lancar / liab. jangka pendek)',
         f'=IFERROR({sum_cats(WTB_J, WTB_C, MAP_AL)}/{sum_cats(WTB_J, WTB_C, MAP_LJPDK, neg=True)},"")',
         '> 1,0', '=IF(B6="","",IF(B6<1,"⚠ di bawah 1","OK"))'),
        ('Debt-to-equity (total liab. / ekuitas)',
         f'=IFERROR({sum_cats(WTB_J, WTB_C, MAP_LIAB_L, neg=True)}/{sum_cats(WTB_J, WTB_C, MAP_EK_LR_L, neg=True)},"")',
         '< 2,0', '=IF(B7="","",IF(OR(B7<0,B7>2),"⚠ leverage/defisit","OK"))'),
        ('Laba (rugi) sebelum pajak',
         '=' + sum_cats(WTB_J, WTB_C, MAP_LR_PRATAX_L, neg=True),
         '> 0', '=IF(B8="","",IF(B8<0,"⚠ rugi","OK"))'),
        ('Interest coverage ((PBT + beban keuangan) / beban keuangan)',
         "=IFERROR((B8+SUMIFS('20_WTB'!$J$6:$J$155,'20_WTB'!$C$6:$C$155,\"Beban Keuangan\"))/SUMIFS('20_WTB'!$J$6:$J$155,'20_WTB'!$C$6:$C$155,\"Beban Keuangan\"),\"\")",
         '> 1,5', '=IF(B9="","",IF(B9<1.5,"⚠ rendah","OK"))'),
    ]
    for i, (lbl, f, bench, ind) in enumerate(rasios):
        r = 6 + i
        cell(ws, f'A{r}', lbl, wrap=True)
        frm(ws, f'B{r}', f, fmt=RP if i == 2 else '0.00', link=True)
        cell(ws, f'C{r}', bench, align='center')
        frm(ws, f'D{r}', ind, bold=True)

    cell(ws, 'A12', 'B. INDIKATOR SA 570 (keuangan · operasi · lainnya)', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 13, ['Indikator', 'Ada?', '', 'Catatan / Rencana Manajemen'])
    indikator = ['Posisi liabilitas neto / modal kerja negatif', 'Pinjaman mendekati jatuh tempo tanpa prospek pembaruan',
                 'Rasio keuangan utama memburuk', 'Kerugian operasi berulang', 'Arus kas operasi negatif',
                 'Tunggakan dividen / pembayaran kreditur', 'Kehilangan manajemen kunci', 'Kehilangan pelanggan/pemasok utama',
                 'Kesulitan tenaga kerja', 'Perkara hukum yang mengancam', 'Perubahan regulasi merugikan', 'Ketidakpastian lain']
    for i, ind in enumerate(indikator):
        r = 14 + i
        cell(ws, f'A{r}', ind, wrap=True)
        inp(ws, f'B{r}')
        inp(ws, f'D{r}')
    dv_name(ws, 'LIST_YATIDAK', 'B14:B25')
    frm(ws, 'B27', '=COUNTIF(B14:B25,"Ya")&" indikator teridentifikasi"', bold=True)
    cell(ws, 'A27', 'Ringkasan:', font=F_BOLD, border=False)
    cell(ws, 'A29', 'KESIMPULAN GOING CONCERN', font=F_BOLD, fill=FILL_SUB, border=False)
    inp(ws, 'A30')
    ws.merge_cells('A30:D31')
    dv_inline(ws, ['Tidak ada keraguan material', 'Keraguan material — pengungkapan MEMADAI (MURTPGC)',
                   'Keraguan material — pengungkapan TIDAK memadai (modifikasi opini)',
                   'Basis kelangsungan usaha tidak tepat (opini tidak wajar)'], 'A30')
    protect(ws)
    return ws


def build_saldoawal(wb):
    ws = wb.create_sheet('32_SaldoAwal')
    title(ws, '32 · Saldo Awal & Auditor Pendahulu (SA 510)', None)
    widths(ws, {'A': 46, 'B': 22, 'C': 22, 'D': 18, 'E': 30})
    info = [('LK tahun lalu diaudit oleh', ''), ('Opini tahun lalu', ''), ('Tanggal laporan auditor pendahulu', '')]
    for i, (lbl, _) in enumerate(info):
        r = 4 + i
        cell(ws, f'A{r}', lbl)
        inp(ws, f'B{r}', fmt=DATE if 'Tanggal' in lbl else None)
        ws.merge_cells(f'B{r}:C{r}')
    cell(ws, 'A9', 'PROSEDUR', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 10, ['Prosedur', 'Status', 'Ref WP', '', 'Catatan'])
    procs = ['Telusur saldo awal ke LK auditan tahun lalu', 'Reviu kertas kerja auditor pendahulu (bila dapat diakses)',
             'Evaluasi konsistensi kebijakan akuntansi', 'Prosedur tambahan atas saldo awal berisiko (persediaan, dll.)']
    for i, p in enumerate(procs):
        r = 11 + i
        cell(ws, f'A{r}', p, wrap=True)
        inp(ws, f'B{r}')
        inp(ws, f'C{r}')
        inp(ws, f'E{r}')
    dv_name(ws, 'LIST_STATUS', 'B11:B14')
    cell(ws, 'A17', 'SELISIH SALDO AWAL (per klien vs LK auditan PY)', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 18, ['Akun', 'Per Klien (awal CY)', 'Per LK Auditan PY', 'Selisih', 'Tindak Lanjut'])
    for r in range(19, 29):
        inp(ws, f'A{r}')
        inp(ws, f'B{r}', fmt=RP)
        inp(ws, f'C{r}', fmt=RP)
        frm(ws, f'D{r}', f'=IF(OR($B{r}="",$C{r}=""),"",$B{r}-$C{r})', fmt=RP)
        inp(ws, f'E{r}')
    protect(ws)
    return ws


def build_subsequent(wb):
    ws = wb.create_sheet('33_PeristiwaSetelah')
    title(ws, '33 · Peristiwa Setelah Periode Pelaporan (SA 560)', 'Cakup s.d. tanggal laporan auditor.')
    widths(ws, {'A': 50, 'B': 14, 'C': 12, 'D': 40})
    cell(ws, 'A4', 'PROSEDUR', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 5, ['Prosedur', 'Status', 'Ref WP', 'Catatan'])
    procs = ['Reviu notulen RUPS/direksi/komisaris setelah tanggal LK', 'Reviu LK interim/manajemen terkini',
             'Tanya manajemen: komitmen, pinjaman baru, litigasi, penjualan aset', 'Reviu jawaban konfirmasi hukum',
             'Reviu transaksi signifikan setelah tanggal neraca (cutoff)', 'Peroleh representasi tertulis s.d. tanggal laporan']
    for i, p in enumerate(procs):
        r = 6 + i
        cell(ws, f'A{r}', p, wrap=True)
        inp(ws, f'B{r}')
        inp(ws, f'C{r}')
        inp(ws, f'D{r}')
    dv_name(ws, 'LIST_STATUS', 'B6:B11')
    cell(ws, 'A14', 'REGISTER PERISTIWA', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 15, ['Uraian Peristiwa', 'Tanggal', 'Tipe', 'Dampak & Perlakuan (penyesuaian / pengungkapan)'])
    for r in range(16, 26):
        inp(ws, f'A{r}')
        inp(ws, f'B{r}', fmt=DATE)
        inp(ws, f'C{r}')
        inp(ws, f'D{r}')
    dv_inline(ws, ['Penyesuai', 'Non-penyesuai'], 'C16:C25')
    protect(ws)
    return ws


def build_related(wb):
    ws = wb.create_sheet('34_PihakBerelasi')
    title(ws, '34 · Pihak Berelasi (SA 550 · PSAK 7)', 'Identifikasi, transaksi, kewajaran syarat, dan kelengkapan pengungkapan.')
    widths(ws, {'A': 26, 'B': 24, 'C': 30, 'D': 16, 'E': 12, 'F': 12, 'G': 10, 'H': 22})
    headrow(ws, 5, ['Pihak Berelasi', 'Sifat Hubungan', 'Sifat Transaksi', 'Nilai Transaksi CY', 'Syarat Wajar (arm\'s length)?', 'Diungkap di LK?', 'Ref WP', 'Flag'])
    for r in range(6, 31):
        for colref in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            inp(ws, f'{colref}{r}', fmt=RP if colref == 'D' else None)
        frm(ws, f'H{r}', f'=IF($A{r}="","",IF($F{r}="Tidak","⚠ belum diungkap",IF($E{r}="Tidak","⚠ evaluasi kewajaran","OK")))', bold=True)
    dv_name(ws, 'LIST_YATIDAK', 'E6:F30')
    ws.freeze_panes = 'A6'
    protect(ws)
    return ws


def build_grupaudit(wb):
    ws = wb.create_sheet('35_GrupAudit')
    title(ws, '35 · Audit Grup (SA 600)', 'Materialitas komponen = % × OM grup (wajib < OM). Lingkup per komponen.')
    widths(ws, {'A': 26, 'B': 18, 'C': 12, 'D': 18, 'E': 20, 'F': 22, 'G': 12, 'H': 26})
    cell(ws, 'A3', 'OM grup:', font=F_BOLD, border=False)
    frm(ws, 'B3', '=MAT_OM', fmt=RP, bold=True, link=True)
    headrow(ws, 5, ['Komponen', 'Total Aset / Signifikansi', '% Mat. Komponen', 'Materialitas Komponen', 'Jenis Pekerjaan', 'Auditor Komponen', 'Status', 'Catatan'])
    for r in range(6, 21):
        inp(ws, f'A{r}')
        inp(ws, f'B{r}', fmt=RP)
        inp(ws, f'C{r}', fmt=PCT0)
        frm(ws, f'D{r}', f'=IF(OR($A{r}="",$C{r}=""),"",ROUND(MIN($C{r},0.99)*MAT_OM,0))', fmt=RP)
        inp(ws, f'E{r}')
        inp(ws, f'F{r}')
        inp(ws, f'G{r}')
        inp(ws, f'H{r}')
    dv_name(ws, 'LIST_KOMPONEN', 'E6:E20')
    dv_name(ws, 'LIST_STATUS', 'G6:G20')
    protect(ws)
    return ws


def build_reliance(wb):
    ws = wb.create_sheet('36_Reliance')
    title(ws, '36 · Mengandalkan Pihak Lain', 'SA 610 (audit internal) · SA 620 (pakar auditor) · SA 402 (organisasi jasa).')
    widths(ws, {'A': 24, 'B': 26, 'C': 14, 'D': 40, 'E': 26})
    blocks = [
        ('A. FUNGSI AUDIT INTERNAL (SA 610)', 'Area kerja yang diandalkan'),
        ('B. PAKAR AUDITOR (SA 620)', 'Bidang keahlian (aktuaria, penilai, hukum, TI)'),
        ('C. ORGANISASI JASA (SA 402)', 'Layanan (payroll, kustodian, cloud) & laporan SOC/3402'),
    ]
    r = 4
    for judul, kol2 in blocks:
        cell(ws, f'A{r}', judul, font=F_BOLD, fill=FILL_SUB, border=False)
        r += 1
        headrow(ws, r, ['Pihak', kol2, 'Kompeten & Objektif?', 'Prosedur Evaluasi Kami', 'Kesimpulan'])
        r += 1
        for rr in range(r, r + 4):
            for colref in ['A', 'B', 'C', 'D', 'E']:
                inp(ws, f'{colref}{rr}')
        dv_name(ws, 'LIST_YATIDAK', f'C{r}:C{r + 3}')
        r += 5
    protect(ws)
    return ws


SAD_TOP, SAD_BOT = 6, 45


def build_sad(wb):
    ws = wb.create_sheet('37_SAD')
    title(ws, '37 · Ikhtisar Salah Saji — SAD Ledger (SA 450)',
          'Dampak laba bertanda: (+) = laba LEBIH saji. Iron curtain = kumulatif neraca; rollover = dampak tahun berjalan.')
    widths(ws, {'A': 6, 'B': 10, 'C': 44, 'D': 13, 'E': 17, 'F': 17, 'G': 16, 'H': 17, 'I': 30})
    headrow(ws, 5, ['No', 'Ref WP', 'Uraian Salah Saji', 'Tipe', 'Dampak Laba CY (±)', 'Dampak Neraca (±)', 'Status Koreksi', 'Dampak Laba dari PY (±)', 'Catatan / Persetujuan Klien'])
    for r in range(SAD_TOP, SAD_BOT + 1):
        for colref in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            inp(ws, f'{colref}{r}', fmt=RP if colref in 'EFH' else None)
    dv_name(ws, 'LIST_TIPE_SAD', f'D{SAD_TOP}:D{SAD_BOT}')
    dv_name(ws, 'LIST_KOREKSI', f'G{SAD_TOP}:G{SAD_BOT}')
    ws['A6'] = 1; ws['B6'] = 'C-5'; ws['C6'] = 'Beban listrik Des belum diakru — DIKOREKSI via AJE-1 (contoh)'
    ws['D6'] = 'Faktual'; ws['E6'] = 25_000_000; ws['F6'] = -25_000_000; ws['G6'] = 'Dikoreksi'

    t = SAD_BOT + 2  # 47
    rows = [
        ('Σ dampak laba DIKOREKSI (cek: sudah jadi AJE Posted?)', f'=SUMIFS(E{SAD_TOP}:E{SAD_BOT},G{SAD_TOP}:G{SAD_BOT},"Dikoreksi")'),
        ('Σ TIDAK dikoreksi — metode IRON CURTAIN (kumulatif)', f'=SUMIFS(E{SAD_TOP}:E{SAD_BOT},G{SAD_TOP}:G{SAD_BOT},"Tidak Dikoreksi")'),
        ('Σ TIDAK dikoreksi — metode ROLLOVER (tahun berjalan)', f'=SUMIFS(E{SAD_TOP}:E{SAD_BOT},G{SAD_TOP}:G{SAD_BOT},"Tidak Dikoreksi")-SUMIFS(H{SAD_TOP}:H{SAD_BOT},G{SAD_TOP}:G{SAD_BOT},"Tidak Dikoreksi")'),
    ]
    for i, (lbl, f) in enumerate(rows):
        cell(ws, f'A{t + i}', lbl, font=F_BOLD, fill=FILL_GREY)
        ws.merge_cells(f'A{t + i}:D{t + i}')
        frm(ws, f'E{t + i}', f, fmt=RP, bold=True, fill=FILL_GREY)
    flags = [
        ('Banding vs OM', f'=IF(MAX(ABS(E{t + 1}),ABS(E{t + 2}))>=MAT_OM,"MATERIAL — koreksi atau modifikasi opini","Tidak material (individu/agregat) vs OM")'),
        ('Banding vs PM', f'=IF(MAX(ABS(E{t + 1}),ABS(E{t + 2}))>=MAT_PM,"≥ PM — ruang agregasi menipis, evaluasi","< PM")'),
    ]
    for i, (lbl, f) in enumerate(flags):
        cell(ws, f'A{t + 3 + i}', lbl, font=F_BOLD)
        frm(ws, f'E{t + 3 + i}', f, bold=True)
        ws.merge_cells(f'E{t + 3 + i}:I{t + 3 + i}')
    note(ws, f'A{t + 6}', 'SA 450: evaluasi baik iron curtain maupun rollover; gunakan yang lebih konservatif. '
                          'Named range SAD_UNCORR (dipakai Cockpit & Opini) = iron curtain.')
    ws.freeze_panes = 'A6'
    protect(ws)
    return ws


def build_evaluasibukti(wb):
    ws = wb.create_sheet('38_EvaluasiBukti')
    title(ws, '38 · Evaluasi Kecukupan & Ketepatan Bukti (SA 500/330)', None)
    widths(ws, {'A': 28, 'B': 24, 'C': 40, 'D': 12, 'E': 12, 'F': 30})
    headrow(ws, 5, ['Area / Akun Signifikan', 'Asersi Utama', 'Bukti Utama yang Diperoleh', 'Cukup?', 'Tepat?', 'Kesimpulan / Tindak Lanjut'])
    for r in range(6, 26):
        for colref in ['A', 'B', 'C', 'D', 'E', 'F']:
            inp(ws, f'{colref}{r}')
    dv_name(ws, 'LIST_ASERSI', 'B6:B25')
    dv_name(ws, 'LIST_YATIDAK', 'D6:E25')
    cell(ws, 'A28', 'KESIMPULAN KESELURUHAN', font=F_BOLD, fill=FILL_SUB, border=False)
    frm(ws, 'B28', '=IF(COUNTIF(D6:E25,"Tidak")>0,"⚠ Ada area dengan bukti belum memadai — selesaikan sebelum opini","Bukti cukup & tepat pada area yang dinilai")', bold=True)
    ws.merge_cells('B28:F28')
    protect(ws)
    return ws
