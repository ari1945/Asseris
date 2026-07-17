"""Helper bersama generator Engagement Pack (styles, proteksi, DV, format)."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation

FONT = 'Arial'
NAVY = '1F2A44'
BLUE = '2563EB'
LINE = 'D5DAE3'
GREY_BG = 'F3F4F6'
INPUT_BG = 'EAF1FB'
ASSUM_BG = 'FFF3C4'
OK_BG = 'E7F6EC'
BAD_BG = 'FDE8E8'
SUB_BG = 'E8ECF3'

RP = '"Rp" #,##0;[Red]("Rp" #,##0);"-"'
NUM = '#,##0;[Red](#,##0);"-"'
PCT = '0.0%;[Red](0.0%);"-"'
PCT0 = '0%'
DATE = 'dd/mm/yyyy'
JAM = '0.0;[Red](0.0);"-"'

F_TITLE = Font(name=FONT, size=14, bold=True, color=NAVY)
F_SUB = Font(name=FONT, size=9, color='6B7280')
F_HEAD = Font(name=FONT, size=9, bold=True, color='FFFFFF')
F_BASE = Font(name=FONT, size=10)
F_BOLD = Font(name=FONT, size=10, bold=True)
F_INPUT = Font(name=FONT, size=10, color='0000FF')
F_LINK = Font(name=FONT, size=10, color='008000')
F_SMALL = Font(name=FONT, size=8, color='6B7280')

FILL_HEAD = PatternFill('solid', start_color=NAVY)
FILL_INPUT = PatternFill('solid', start_color=INPUT_BG)
FILL_ASSUM = PatternFill('solid', start_color=ASSUM_BG)
FILL_SUB = PatternFill('solid', start_color=SUB_BG)
FILL_GREY = PatternFill('solid', start_color=GREY_BG)

THIN = Side(style='thin', color=LINE)
B_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
B_BOT = Border(bottom=THIN)

UNLOCK = Protection(locked=False)


def title(ws, text, sub=None):
    ws['A1'] = text
    ws['A1'].font = F_TITLE
    if sub:
        ws['A2'] = sub
        ws['A2'].font = F_SUB
    ws.sheet_view.showGridLines = False


def headrow(ws, row, cols, start=1, height=24):
    for i, label in enumerate(cols):
        c = ws.cell(row=row, column=start + i, value=label)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = B_ALL
    ws.row_dimensions[row].height = height


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def cell(ws, ref, value=None, font=None, fmt=None, fill=None, border=True, align=None, wrap=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = font or F_BASE
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = B_ALL
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c


def inp(ws, ref, value=None, fmt=None, assum=False):
    """Sel input: teks biru, fill khas, TIDAK terkunci."""
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = F_INPUT
    c.fill = FILL_ASSUM if assum else FILL_INPUT
    c.border = B_ALL
    c.protection = UNLOCK
    if fmt:
        c.number_format = fmt
    return c


def frm(ws, ref, formula, fmt=None, bold=False, fill=None, link=False):
    c = ws[ref]
    c.value = formula
    c.font = F_BOLD if bold else (F_LINK if link else F_BASE)
    c.border = B_ALL
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


def dv_name(ws, list_name, rng, blank=True):
    dv = DataValidation(type='list', formula1=f'={list_name}', allow_blank=blank, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(rng)
    return dv


def unlock_range(ws, rng, fmt=None, input_style=True):
    for row in ws[rng]:
        for c in row:
            c.protection = UNLOCK
            if input_style:
                c.font = F_INPUT
                c.fill = FILL_INPUT
            c.border = B_ALL
            if fmt:
                c.number_format = fmt


def border_range(ws, rng):
    for row in ws[rng]:
        for c in row:
            c.border = B_ALL
            if c.font is None or c.font.name != FONT:
                c.font = F_BASE


def protect(ws):
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


def note(ws, ref, text):
    c = ws[ref]
    c.value = text
    c.font = F_SMALL
    c.alignment = Alignment(wrap_text=True, vertical='top')
