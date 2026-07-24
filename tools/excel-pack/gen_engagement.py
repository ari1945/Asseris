"""Ruang Kerja Perikatan: 00_Cockpit, 01_Tugas, 02_Program, 03_CatatanReviu, 04_TimeBudget, 05_Timeline."""
from xhelp import *
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

FASES = ['Perencanaan', 'Pekerjaan Lapangan', 'Penyelesaian', 'Pelaporan']

PROGRAM = [
    # (fase, prosedur, SA, ref WP)
    ('Perencanaan', 'Evaluasi penerimaan/keberlanjutan klien & independensi tim', 'SA 210 · 220', 'A-1/A-3'),
    ('Perencanaan', 'Terbitkan/perbarui surat perikatan', 'SA 210', 'A-2'),
    ('Perencanaan', 'Pahami entitas, industri & lingkungannya', 'SA 315', 'A-6'),
    ('Perencanaan', 'Tetapkan materialitas OM/PM/CTT', 'SA 320', 'A-5'),
    ('Perencanaan', 'Identifikasi & nilai RoMM per asersi (register risiko)', 'SA 315', 'A-7/A-8'),
    ('Perencanaan', 'Pahami pengendalian internal & ITGC; evaluasi D&I', 'SA 315', 'A-9'),
    ('Perencanaan', 'Prosedur analitis awal', 'SA 520', 'A-10'),
    ('Perencanaan', 'Susun strategi & program audit; diskusi tim (fraud brainstorming)', 'SA 300 · 240', 'A-4/B-1'),
    ('Pekerjaan Lapangan', 'Uji pengendalian (TOC) atas siklus yang diandalkan', 'SA 330', 'B-2'),
    ('Pekerjaan Lapangan', 'Kas & bank: rekonsiliasi + konfirmasi bank', 'SA 505', 'C-1'),
    ('Pekerjaan Lapangan', 'Piutang: konfirmasi + penerimaan setelah tanggal neraca', 'SA 505', 'C-2'),
    ('Pekerjaan Lapangan', 'Pendapatan: cutoff 2 sisi & vouching ke dokumen serah terima', 'SA 240', 'C-2'),
    ('Pekerjaan Lapangan', 'Persediaan: observasi stock opname + uji NRV', 'SA 501', 'C-3'),
    ('Pekerjaan Lapangan', 'Aset tetap: vouching penambahan + reviu penyusutan & impairment', 'PSAK 16/48', 'C-4'),
    ('Pekerjaan Lapangan', 'Utang: search for unrecorded liabilities + konfirmasi', 'SA 505', 'C-5'),
    ('Pekerjaan Lapangan', 'Pinjaman: konfirmasi + uji kepatuhan covenant', '—', 'C-6'),
    ('Pekerjaan Lapangan', 'Ekuitas: telusur akta & RUPS', '—', 'C-7'),
    ('Pekerjaan Lapangan', 'Payroll & imbalan kerja: uji perhitungan + aktuaria bila ada', 'PSAK 24 · SA 620', 'C-8'),
    ('Pekerjaan Lapangan', 'Pajak: rekonsiliasi book-tax + reviu SPT & sengketa', 'PSAK 46', 'C-9'),
    ('Pekerjaan Lapangan', 'Journal entry testing (Benford, akhir pekan, near-CTT)', 'SA 240', 'B-4'),
    ('Pekerjaan Lapangan', 'Sampling MUS atas populasi terpilih', 'SA 530', 'B-3'),
    ('Pekerjaan Lapangan', 'Pihak berelasi: identifikasi, konfirmasi & kewajaran', 'SA 550', 'C-10'),
    ('Pekerjaan Lapangan', 'Estimasi akuntansi: evaluasi asumsi manajemen', 'SA 540', '—'),
    ('Penyelesaian', 'Prosedur analitis akhir', 'SA 520', 'D-2'),
    ('Penyelesaian', 'Akumulasi & evaluasi salah saji (SAD) vs materialitas', 'SA 450', 'D-1'),
    ('Penyelesaian', 'Evaluasi going concern & rencana manajemen', 'SA 570', 'D-3'),
    ('Penyelesaian', 'Reviu peristiwa setelah periode pelaporan', 'SA 560', 'D-4'),
    ('Penyelesaian', 'Peroleh surat representasi manajemen', 'SA 580', 'D-5'),
    ('Penyelesaian', 'Evaluasi kecukupan & ketepatan bukti', 'SA 500', 'D-6'),
    ('Penyelesaian', 'Daftar-uji pengungkapan LK', 'PSAK', 'D-7'),
    ('Pelaporan', 'Penelaahan mutu perikatan (EQR) bila dipersyaratkan', 'ISQM 2', 'D-8'),
    ('Pelaporan', 'Rumuskan opini & susun laporan auditor', 'SA 700/705', 'D-9'),
    ('Pelaporan', 'Komunikasi TCWG & management letter', 'SA 260/265', 'D-10'),
    ('Pelaporan', 'Rakit arsip final ≤ 60 hari', 'SA 230', 'D-11'),
]
PRG_TOP, PRG_BOT = 6, 65


def build_cockpit(wb):
    ws = wb.create_sheet('00_Cockpit')
    title(ws, '00 · Engagement Cockpit', 'Seluruh angka pada sheet ini turunan otomatis — tidak ada input.')
    widths(ws, {'A': 34, 'B': 22, 'C': 6, 'D': 34, 'E': 22})

    cell(ws, 'A4', 'PERIKATAN', font=F_BOLD, fill=FILL_SUB, border=False)
    info = [('Klien', '=KLIEN'), ('Tanggal LK', '=TEXT(PERIODE_AKHIR,"dd/mm/yyyy")'),
            ('No Perikatan', "='SETUP'!B12"), ('Rekan Perikatan', "='SETUP'!B17")]
    for i, (lbl, f) in enumerate(info):
        cell(ws, f'A{5 + i}', lbl)
        frm(ws, f'B{5 + i}', f, bold=True, link=True)

    cell(ws, 'D4', 'MATERIALITAS (SA 320)', font=F_BOLD, fill=FILL_SUB, border=False)
    for i, (lbl, nm) in enumerate([('OM', 'MAT_OM'), ('PM', 'MAT_PM'), ('CTT', 'MAT_CTT')]):
        cell(ws, f'D{5 + i}', lbl)
        frm(ws, f'E{5 + i}', f'={nm}', fmt=RP, bold=True, link=True)
    cell(ws, 'D8', 'Status WTB')
    frm(ws, 'E8', "='20_WTB'!K3", bold=True, link=True)

    cell(ws, 'A11', 'PROGRES PROGRAM AUDIT (02_Program)', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 12, ['Fase', 'Selesai', '', 'Total', '% '])
    for i, fase in enumerate(FASES):
        r = 13 + i
        cell(ws, f'A{r}', fase)
        frm(ws, f'B{r}', f'=COUNTIFS(\'02_Program\'!$B${PRG_TOP}:$B${PRG_BOT},A{r},\'02_Program\'!$G${PRG_TOP}:$G${PRG_BOT},"Selesai")', fmt=NUM, link=True)
        frm(ws, f'D{r}', f'=COUNTIFS(\'02_Program\'!$B${PRG_TOP}:$B${PRG_BOT},A{r})', fmt=NUM, link=True)
        frm(ws, f'E{r}', f'=IF(D{r}=0,"",B{r}/D{r})', fmt=PCT)

    cell(ws, 'A18', 'PROGRES KERTAS KERJA (22_IndeksWP)', font=F_BOLD, fill=FILL_SUB, border=False)
    headrow(ws, 19, ['Fase', 'Final', '', 'Total', '% '])
    for i, fase in enumerate(FASES):
        r = 20 + i
        cell(ws, f'A{r}', fase)
        frm(ws, f'B{r}', f'=COUNTIFS(\'22_IndeksWP\'!$C$6:$C$55,A{r},\'22_IndeksWP\'!$J$6:$J$55,"Final")', fmt=NUM, link=True)
        frm(ws, f'D{r}', f'=COUNTIFS(\'22_IndeksWP\'!$C$6:$C$55,A{r})', fmt=NUM, link=True)
        frm(ws, f'E{r}', f'=IF(D{r}=0,"",B{r}/D{r})', fmt=PCT)

    cell(ws, 'A25', 'INDIKATOR KUNCI', font=F_BOLD, fill=FILL_SUB, border=False)
    kpis = [
        ('Risiko signifikan teridentifikasi', "='10_Risiko'!D37"),
        ('Catatan reviu masih OPEN', '=COUNTIF(\'03_CatatanReviu\'!$H$6:$H$55,"Open")'),
        ('Tugas OPEN / terlambat', '=COUNTIF(\'01_Tugas\'!$F$6:$F$45,"Open")&" / "&COUNTIF(\'01_Tugas\'!$G$6:$G$45,"TERLAMBAT")'),
        ('AJE posted', '=COUNTIF(\'21_AJE\'!$G$6:$G$65,"Posted")'),
        ('Konfirmasi belum kembali', '=COUNTIF(\'30_Konfirmasi\'!$I$6:$I$45,"Tidak Kembali")'),
        ('Analitis: akun flag INVESTIGASI', '=COUNTIF(\'24_Analitis\'!$G$6:$G$155,"INVESTIGASI")'),
    ]
    for i, (lbl, f) in enumerate(kpis):
        cell(ws, f'A{26 + i}', lbl)
        frm(ws, f'B{26 + i}', f, bold=True, link=True)

    cell(ws, 'D25', 'GERBANG PELAPORAN (SA 450 · 700)', font=F_BOLD, fill=FILL_SUB, border=False)
    gates = [
        ('Σ salah saji TIDAK dikoreksi (laba)', '=SAD_UNCORR', RP),
        ('vs OM', '=IF(ABS(SAD_UNCORR)>=MAT_OM,"≥ OM — MATERIAL","< OM")', None),
        ('vs PM (agregasi mendekati?)', '=IF(ABS(SAD_UNCORR)>=MAT_PM,"≥ PM — waspada agregasi","< PM")', None),
        ('Opini indikatif (42_Opini)', "='42_Opini'!D5", None),
    ]
    for i, (lbl, f, fmt_) in enumerate(gates):
        cell(ws, f'D{26 + i}', lbl)
        frm(ws, f'E{26 + i}', f, fmt=fmt_, bold=True, link=True)
    protect(ws)
    return ws


def build_tugas(wb):
    ws = wb.create_sheet('01_Tugas')
    title(ws, '01 · Daftar Tugas Tim', 'Aging otomatis vs TODAY(). Tanpa isolasi per-user (keterbatasan Excel).')
    widths(ws, {'A': 8, 'B': 46, 'C': 16, 'D': 10, 'E': 12, 'F': 10, 'G': 14, 'H': 30})
    headrow(ws, 5, ['No', 'Tugas', 'PIC', 'Ref WP', 'Jatuh Tempo', 'Status', 'Aging', 'Catatan'])
    for r in range(6, 46):
        for colref in ['A', 'B', 'C', 'D', 'E', 'F', 'H']:
            inp(ws, f'{colref}{r}', fmt=DATE if colref == 'E' else None)
        frm(ws, f'G{r}', f'=IF(OR($E{r}="",$F{r}="Selesai"),"",IF(TODAY()>$E{r},"TERLAMBAT",$E{r}-TODAY()&" hari"))', bold=True)
    dv_name(ws, 'TIM', 'C6:C45')
    dv_name(ws, 'LIST_TASK', 'F6:F45')
    ws['A6'] = 1; ws['B6'] = 'Kirim daftar permintaan data (PBC) ke klien (contoh)'
    ws['F6'] = 'Open'
    ws.freeze_panes = 'A6'
    protect(ws)
    return ws


def build_program(wb):
    ws = wb.create_sheet('02_Program')
    title(ws, '02 · Program Audit', 'Program baku turunan metodologi Asseris — sesuaikan dengan hasil penilaian risiko (SA 330).')
    widths(ws, {'A': 6, 'B': 20, 'C': 62, 'D': 16, 'E': 10, 'F': 14, 'G': 11, 'H': 12, 'I': 30})
    headrow(ws, 5, ['No', 'Fase', 'Prosedur', 'Standar', 'Ref WP', 'PIC', 'Status', 'Tanggal', 'Catatan'])
    for i, (fase, proc, sa, ref) in enumerate(PROGRAM):
        r = PRG_TOP + i
        cell(ws, f'A{r}', i + 1, align='center')
        cell(ws, f'B{r}', fase)
        cell(ws, f'C{r}', proc, wrap=True)
        cell(ws, f'D{r}', sa, font=F_SMALL)
        cell(ws, f'E{r}', ref, align='center')
        inp(ws, f'F{r}')
        inp(ws, f'G{r}')
        inp(ws, f'H{r}', fmt=DATE)
        inp(ws, f'I{r}')
    for r in range(PRG_TOP + len(PROGRAM), PRG_BOT + 1):
        for colref in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            inp(ws, f'{colref}{r}', fmt=DATE if colref == 'H' else None)
    dv_name(ws, 'LIST_FASE', f'B{PRG_TOP}:B{PRG_BOT}')
    dv_name(ws, 'TIM', f'F{PRG_TOP}:F{PRG_BOT}')
    dv_name(ws, 'LIST_STATUS', f'G{PRG_TOP}:G{PRG_BOT}')
    ws.freeze_panes = 'A6'
    protect(ws)
    return ws


def build_catatanreviu(wb):
    ws = wb.create_sheet('03_CatatanReviu')
    title(ws, '03 · Catatan Reviu', 'Reviewer menulis catatan; preparer merespons; reviewer menutup. Open = belum tuntas.')
    widths(ws, {'A': 8, 'B': 12, 'C': 14, 'D': 14, 'E': 10, 'F': 44, 'G': 44, 'H': 10, 'I': 12})
    cell(ws, 'A3', 'Masih OPEN:', font=F_BOLD, border=False)
    frm(ws, 'B3', '=COUNTIF(H6:H55,"Open")', fmt=NUM, bold=True)
    headrow(ws, 5, ['No', 'Tanggal', 'Dari', 'Kepada', 'Ref WP', 'Catatan Reviu', 'Respons', 'Status', 'Tgl Tutup'])
    for r in range(6, 56):
        for colref in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            inp(ws, f'{colref}{r}', fmt=DATE if colref in 'BI' else None)
    dv_name(ws, 'TIM', 'C6:C55')
    dv_name(ws, 'TIM', 'D6:D55')
    dv_name(ws, 'LIST_TASK', 'H6:H55')
    ws.freeze_panes = 'A6'
    protect(ws)
    return ws


def build_timebudget(wb):
    ws = wb.create_sheet('04_TimeBudget')
    title(ws, '04 · Time & Budget', 'Jam & tarif per anggota tim; varians dan realisasi fee otomatis.')
    widths(ws, {'A': 28, 'B': 16, 'C': 14, 'D': 14, 'E': 14, 'F': 18, 'G': 18})
    headrow(ws, 5, ['Anggota Tim', 'Tarif / Jam', 'Jam Anggaran', 'Jam Aktual', 'Varians Jam', 'Biaya Anggaran', 'Biaya Aktual'])
    for i in range(6):
        r = 6 + i
        frm(ws, f'A{r}', f"=IF('SETUP'!B{17 + i}=\"\",\"\",'SETUP'!B{17 + i})", link=True)
        inp(ws, f'B{r}', fmt=RP)
        inp(ws, f'C{r}', fmt=JAM)
        inp(ws, f'D{r}', fmt=JAM)
        frm(ws, f'E{r}', f'=IF($A{r}="","",N(C{r})-N(D{r}))', fmt=JAM)
        frm(ws, f'F{r}', f'=IF($A{r}="","",N(B{r})*N(C{r}))', fmt=RP)
        frm(ws, f'G{r}', f'=IF($A{r}="","",N(B{r})*N(D{r}))', fmt=RP)
    cell(ws, 'A12', 'TOTAL', font=F_BOLD, fill=FILL_GREY)
    for colref in ['C', 'D', 'F', 'G']:
        frm(ws, f'{colref}12', f'=SUM({colref}6:{colref}11)', fmt=RP if colref in 'FG' else JAM, bold=True, fill=FILL_GREY)
    rows = [('Fee perikatan (kontrak)', None, 'inp'),
            ('Realisasi biaya vs fee', '=IF(N(B14)=0,"",G12/B14)', 'frm'),
            ('Margin (fee − biaya aktual)', '=IF(N(B14)=0,"",B14-G12)', 'frm')]
    for i, (lbl, f, kind) in enumerate(rows):
        r = 14 + i
        cell(ws, f'A{r}', lbl, font=F_BOLD)
        if kind == 'inp':
            inp(ws, f'B{r}', fmt=RP)
        else:
            frm(ws, f'B{r}', f, fmt=PCT if 'Realisasi' in lbl else RP, bold=True)
    protect(ws)
    return ws


def build_timeline(wb):
    ws = wb.create_sheet('05_Timeline')
    title(ws, '05 · Jadwal & Lini Masa Audit', 'Isi mulai/selesai per aktivitas — bar Gantt mingguan otomatis (16 minggu dari tanggal grid).')
    widths(ws, {'A': 34, 'B': 12, 'C': 12})
    for i in range(16):
        ws.column_dimensions[chr(ord('D') + i) if i < 23 else 'Z'].width = 5
    cell(ws, 'A3', 'Awal grid (minggu-1):', font=F_BOLD, border=False)
    inp(ws, 'B3', None, fmt=DATE, assum=True)
    ws['B3'] = '=PERIODE_AKHIR-70'
    ws['B3'].font = F_INPUT
    headrow(ws, 5, ['Aktivitas / Tonggak', 'Mulai', 'Selesai'] + [f'M{i + 1}' for i in range(16)])
    for i in range(16):
        col = chr(ord('D') + i)
        frm(ws, f'{col}4', f'=TEXT($B$3+{i * 7},"dd/mm")', fill=FILL_GREY)
    prefill = ['Perencanaan & penilaian risiko', 'Interim / uji pengendalian', 'Observasi stock opname',
               'Kirim konfirmasi', 'Pekerjaan lapangan final', 'Reviu manajer', 'Reviu partner & EQR',
               'Draf LK & laporan auditor', 'Perakitan arsip final']
    for i in range(20):
        r = 6 + i
        inp(ws, f'A{r}', prefill[i] if i < len(prefill) else None)
        inp(ws, f'B{r}', fmt=DATE)
        inp(ws, f'C{r}', fmt=DATE)
        for j in range(16):
            col = chr(ord('D') + j)
            c = ws[f'{col}{r}']
            c.border = B_ALL
    bar = PatternFill('solid', start_color='2563EB')
    ws.conditional_formatting.add('D6:S25',
        FormulaRule(formula=['AND($B6<>"",$C6<>"",D$4<>"",$B6<=($B$3+(COLUMN(D$4)-COLUMN($D$4))*7+6),$C6>=($B$3+(COLUMN(D$4)-COLUMN($D$4))*7))'], fill=bar))
    protect(ws)
    return ws
